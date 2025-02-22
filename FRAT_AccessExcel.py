from openpyxl import Workbook
from datetime import datetime
from openpyxl import load_workbook

def createTodaysFile(fname, studname):
    print("**********Creating Today's file *****************", fname)
        
    #creating an excel file
    wb=Workbook()

    #creating a worksheet
    sheet = wb.active

    #get current time
    at_time = datetime.now()

    #convert time to string of the given pattern
    current_time= at_time.strftime ('%H:%M:%S')
    
    id=1
    sheet[f'A{id}'] = studname
    sheet[f'B{id}'] = current_time

    wb.save (fname)
    wb.close()

def updateAtt(fname, studname):
    print("**********Updating Today's file *****************")

    '''print("File created is :", fname)
    print("studname = ", studname)
      '''
    #flag to check new/duplicate entry
    new_entry=True
    #loading the excel file
    wb=load_workbook(fname)

    #loading the worksheet
    sheet = wb['Sheet']

    entries=[cell.value for cell in sheet['A']]
    print("names in excel :", entries)

    for name in entries:
        if name == studname:
            print("you have already logged in")
            new_entry=False
            break
        else:
            print("new entry")
            continue

    if (new_entry == True):
        #get current time
        at_time = datetime.now()

        #convert time to string of the given pattern
        current_time= at_time.strftime ('%H:%M:%S')
        print("current_time : ", current_time)

        #Find the next row to be filled
        id=sheet.max_row + 1
        print("row filled is : ", id)

        #for id, record in enumerate(xldata):
        sheet[f'A{id}'] = studname
        sheet[f'B{id}'] = current_time
        wb.save (fname)
    wb.close()

    



